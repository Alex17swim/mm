# plt_excel.py
# chengyu wang, 2021-10-14, XJTLU, last update: 2021-10-14 18:58
import os, sys
import numpy as np
# import torch
import matplotlib.pyplot as plt
import openpyxl, csv

_import_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(_import_root)
from all_include.pt_head import *
from all_include.pt_utils_file import get_fns
from all_include.pt_metric import get_confusion_matrix, plt_confusion_matrix, cm_axis_name, show_confMat

LOC = ['best', 'upper right', 'upper left', 'lower left', 'lower right', 'right',
       'center left', 'center right', 'lower center', 'upper center', 'center']
FONT_SIZE = ['xx-large', 'x-large', 'larger', 'large', 'xx-small', 'x-small', 'smaller', 'small', 'medium', 'None']
COLORS = ['g', 'b', 'm', 'y', 'c', 'r', 'k',
          'lime', 'cornflowerblue', 'plum', 'yellow', 'cyan', 'brown', 'gray',
          'lawngreen', 'deepskyblue', 'purple', 'gold', 'darkcyan', 'tomato', 'silver',
          'green', 'blue', 'violet', 'orange', 'paleturquoise', 'red', 'black']
LINESTYLES = ['-', '--', '-.', ':']
MARKER = ['o', '*', 'x', 'd', '<','>', '+', 's', '.', 'p', 'h']


def make_matrix(_fnp_xls, _st_num=0, _fp_out=None, _num_cls=24, _num_cell=46, _log_cm=None, _use_bg=False):
    assert osp.exists(_fnp_xls), _fnp_xls
    xls = openpyxl.load_workbook(_fnp_xls)
    for _st_idx in range(_st_num):
        st = xls.worksheets[_st_idx]
        # for st in xls.worksheets:
        print("going to read {} with {} rows and {}({}) columns".format(st.title, st.max_row, st.max_column, _num_cell))
        # if st.max_column != _num_cell:
        #     r = input("! expect {} ({}) chromosomes, input 'c' to continue".format(_num_cell, st.max_column))
        #     if r not in ['c', 'C']: raise ValueError("886")

        if _fp_out is None: _out_dir =\
            osp.join(TEMP_ROOT, 'log', "{}_{}_cm_{}".format(osp.splitext(osp.basename(_fnp_xls))[0], st.title, time_str()))
        else: _out_dir = _fp_out
        _log_cm = osp.join(_out_dir, 'log.txt')
        os.makedirs(_out_dir, exist_ok=True)

        pred_list, lb_list, lb46 = [], [], []
        for i in range(0, _num_cls-2): lb46.append(i); lb46.append(i)
        lb46.append(_num_cls-2); lb46.append(_num_cls-1) # X, Y
        _range = list(np.arange(_num_cls, dtype=int))  # <=23
        if _use_bg:
            lb46.append(_num_cls) # (_use_bg-1) # _use_bg = 24+1
            _range.append(_num_cls) # (_use_bg-1) # _use_bg = 24+1
        for i, row in enumerate(st.iter_rows()):
            r2 = []
            for cell in row:
                r2.append(cell.value)
            r2 = r2[:_num_cell] # only check 46 columns
            if None in r2:
                write_txt([_log_cm], "!!delete the No.{} row: {}".format(i+1, r2))
                continue
            r22 = []
            # r22 = [int(r-1) for r in r2]
            for r in r2:
                if r <= _num_cls: r22.append(int(r-1))
                else: r22.append(int(_num_cls))
            pred_list.append(r22)
            lb_list.append(lb46)
            # for cell in row:
        _num_case, matrix85 = len(pred_list), np.zeros((len(_range), len(_range)))
        confusion_mat = np.zeros((_num_case, len(_range), len(_range)))

        wrong85, total85, bg85, i = [], [], [], -1
        assert len(lb_list) == _num_case
        for i in range(0, _num_case):
            np_lb, pred = lb_list[i], pred_list[i]
            confusion_mat[i], _wrg, _ttl, _bg = get_confusion_matrix(_num_cls, np_lb, pred, _log_list=_log_cm, _use_bg=_use_bg)
            wrong85.append(_wrg); total85.append(_ttl); bg85.append(_bg)
            matrix85 += confusion_mat[i]
        ##_matrix_N = matrix85.copy()
        # for i in _range: # range(_num_cls):
        #     # i = idx-1
        #     c = matrix85[i, :]
        #     wrong85.append(int(c.sum() - c[i]))
        #     total85.append(int(c.sum()))
        ##     _matrix_N[i, :] = c[i] / c.sum() if c.sum() > 0 else 0
        #     bg85.append(int(c[len(_range)-1]))
        #     # write_txt(_log_cm, "No.{} wrong=[{}/{}]".format(i + 1, wrong85[i], total85[i]), _prt)
        write_txt(_log_cm, " wrong=({}){}\n total({})={}\n bg=({}){}".
                  format(sum(wrong85), wrong85, sum(total85), total85, int(sum(bg85)), bg85))
        wrong85, total85, bg85 = int(sum(wrong85)), int(sum(total85)), int(sum(bg85))
        plt_confusion_matrix(matrix85, cm_axis_name, 'final_all_{}samples_{}err_{}bg'
                             .format(total85, wrong85, bg85), _out_dir, wrong85, total85, bg85)
        _recall_list, _preci_list, _correct_list, _acc, _recall, _preci = \
            show_confMat(matrix85, cm_axis_name, 'mchr', 0, _out_dir=None, _log=_log_cm)
        write_txt(_log_cm, "sheet={}, _acc={}, _recall={}, _preci={}\n output={}"
                  .format(st.title, _acc, _recall, _preci, _out_dir))
# end make_matrix

def from_excel(_fp_in, _fp_out=None, _ignore_list=None):
    assert osp.exists(_fp_in), "!check input:{}".format(_fp_in)
    _fns, _dirs, _files, _ld, _lf, _ldf = get_fns(_fp_in)
    assert _lf == _ldf

    print("loaded {} '{}' files from: {}".format(_lf, osp.splitext(_fns[0])[1], _fp_in))
    for file in _files:
        xls = openpyxl.load_workbook(osp.join(_fp_in, file))
        ws = xls.worksheets
        print("-loading: {}".format(file))
# end from_excel

def from_csv(_fp_in, _fp_out=None, _tag='from csv', _ext='.pdf'):
    assert osp.exists(_fp_in), "!check input:{}".format(_fp_in)
    _fns, _dirs, _files, _ld, _lf, _ldf = get_fns(_fp_in)
    assert _lf == _ldf

    if _fp_out is None: # osp.join(osp.dirname(_fp_in), "{}_csv_{}".format(osp.basename(_fp_in), time_str()))
        _fp_out = osp.join(str(_fp_in).rstrip(RSTRIP) + "_csv_{}".format(time_str()))
    os.makedirs(_fp_out, exist_ok=True)
    _log = osp.join(_fp_out, 'log_{}.txt'.format(time_str()))
    write_txt([_log], "loaded {} '{}' files from: {}".format(_lf, osp.splitext(_fns[0])[1], _fp_in))
    value_list = []; num_list = []; nf = 0
    legend_list = []
    for i, file in enumerate(_files):
        _fnp = osp.join(_fp_in, file)
        legend_list.append(file[: file.index('.')])
        _n_list, _v_list = [], []
        with open(_fnp) as f:
            r = csv.reader(f)
            print("head: {}".format(next(r)))
            for row in r:
                _n_list.append(int(row[1]))
                _v_list.append(round(float(row[2]), PRINT_DECIMAL))
        write_txt([_log], "-loading: {}".format(len(_n_list)))
        value_list.append(_v_list)
        if num_list == [] or len(num_list[0]) != len(_n_list):
            num_list.append(_n_list)
    dim = len(value_list[0])
    write_txt([_log], "- from_csv done with {} files, dim={}".format(i, dim))

    _fnp = osp.join(_fp_out, "{}{}".format(_tag, _ext))
    if _ext != '.jpg':
        _fs = 18
        plt.rcParams['figure.figsize'] = (8, 6) # fig = plt.figure(_tag, figsize=(8, 6), clear=True)
    else:
        _fs = 22
        plt.rcParams['figure.figsize'] = (16, 14) # fig = plt.figure(_tag, figsize=(16, 14), clear=True)
    _fs2 = _fs - 6

    xlocations = np.array(range(_lf))
    x_span = np.arange(1, 1 + dim, 1)
    # x_name_list = list(range(_lf))
    # y_name_list = x_name_list # list(range(1.0))
    # plt.xticks(xlocations, x_name_list) # , rotation=60)
    # plt.yticks(xlocations, y_name_list)
    plt.locator_params(axis='x', nbins=5)
    plt.locator_params(axis='y', nbins=10)
    # plt.xlim((1, dim))  # ! this will jump figure window
    plt.ylim((0.1, 1.05))  # ! this will jump figure window

    plt.xlabel('epoch', fontsize=_fs)
    plt.ylabel('True label', fontsize=_fs)
    plt.tick_params(labelsize=_fs2)
    plt.grid(axis='both', alpha=0.5)
    # if _ext == '.jpg':
    #     plt.title("{}_of_{}_class".format(_tag, _lf), fontsize=26, y=1.02)  # 'xx-large'

    for i in range(len(value_list)): # marker=MARKER[i],
        plt.plot(x_span, value_list[i], linestyle=LINESTYLES[i % 2], linewidth=1, color=COLORS[i]) # , label=legend_list[_i])

    plt.legend(legend_list, fontsize=_fs2, loc='lower right' ) # 'best' 'lower right'
    # plt.legend(['1', '2', '3', '4', '5', '6', '7', '8'])

    if _ext != '.jpg':
        plt.subplots_adjust(top=0.98, bottom=0.12, right=0.9, left=0.1, hspace=0, wspace=0)
        plt.margins(0, 0)
        plt.savefig(_fnp[:-4] + '.jpg')
        plt.pause(1)  # plt.show()
        fig.savefig(_fnp, dpi=200)
        plt.close()
    else:
        plt.savefig(_fnp)
    plt.cla() # 0404
# end from_csv

def from_txt(_fp_in, _fp_out=None, _tag='from txt', _ext='.pdf'):
    assert osp.exists(_fp_in), "!check input:{}".format(_fp_in)
    _fns, _dirs, _files, _ld, _lf, _ldf = get_fns(_fp_in)
    assert _lf == _ldf

    if _fp_out is None: # osp.join(osp.dirname(_fp_in), "{}_csv_{}".format(osp.basename(_fp_in), time_str()))
        _fp_out = osp.join(str(_fp_in).rstrip(RSTRIP) + "_csv_{}".format(time_str()))
    os.makedirs(_fp_out, exist_ok=True)
    _log = osp.join(_fp_out, 'log_{}.txt'.format(time_str()))

    write_txt([_log], "loaded {} '{}' files from: {}".format(_lf, osp.splitext(_fns[0])[1], _fp_in))
    value_list = []; num_list = []; nf = 0
    legend_list = []
    for i, file in enumerate(_files):
        _fnp = osp.join(_fp_in, file)
        legend_list.append(file[: file.index('.')])
        _n_list, _v_list = [], []
        with open(_fnp) as f:
            r = f.readline()
            print("head: {}".format(next(r)))
            for row in r:
                _n_list.append(int(row[1]))
                _v_list.append(round(float(row[2]), PRINT_DECIMAL))
        write_txt([_log], "-loading: {}".format(len(_n_list)))
        value_list.append(_v_list)
        if num_list == [] or len(num_list[0]) != len(_n_list):
            num_list.append(_n_list)
    dim = len(value_list[0])
    write_txt([_log], "- from_csv done with {} files, dim={}".format(i, dim))

    _fnp = osp.join(_fp_out, "{}{}".format(_tag, _ext))
    if _ext != '.jpg':
        _fs = 18
        plt.rcParams['figure.figsize'] = (8, 6) # fig = plt.figure(_tag, figsize=(8, 6), clear=True)
    else:
        _fs = 22
        plt.rcParams['figure.figsize'] = (16, 14) # fig = plt.figure(_tag, figsize=(16, 14), clear=True)
    _fs2 = _fs - 6

    xlocations = np.array(range(_lf))
    x_span = np.arange(1, 1 + dim, 1)
    # x_name_list = list(range(_lf))
    # y_name_list = x_name_list # list(range(1.0))
    # plt.xticks(xlocations, x_name_list) # , rotation=60)
    # plt.yticks(xlocations, y_name_list)
    plt.locator_params(axis='x', nbins=5)
    plt.locator_params(axis='y', nbins=10)
    # plt.xlim((1, dim))  # ! this will jump figure window
    plt.ylim((0.1, 1.05))  # ! this will jump figure window

    plt.xlabel('epoch', fontsize=_fs)
    plt.ylabel('True label', fontsize=_fs)
    plt.tick_params(labelsize=_fs2)
    plt.grid(axis='both', alpha=0.5)
    # if _ext == '.jpg':
    #     plt.title("{}_of_{}_class".format(_tag, _lf), fontsize=26, y=1.02)  # 'xx-large'

    for i in range(len(value_list)): # marker=MARKER[i],
        plt.plot(x_span, value_list[i], linestyle=LINESTYLES[i % 2], linewidth=1, color=COLORS[i]) # , label=legend_list[_i])

    plt.legend(legend_list, fontsize=_fs2, loc='lower right' ) # 'best' 'lower right'
    # plt.legend(['1', '2', '3', '4', '5', '6', '7', '8'])

    if _ext != '.jpg':
        plt.subplots_adjust(top=0.98, bottom=0.12, right=0.9, left=0.1, hspace=0, wspace=0)
        plt.margins(0, 0)
        plt.savefig(_fnp[:-4] + '.jpg')
        plt.pause(1)  # plt.show()
        fig.savefig(_fnp, dpi=200)
        plt.close()
    else:
        plt.savefig(_fnp)
# end from_txt


if __name__ == '__main__':
    # make_matrix(osp.join(TEMP_ROOT, 'excel', 'tmi', 'f2.xlsx'), _st_num=2, _fp_out=None, _use_bg=True)
    from_csv(osp.join(TEMP_ROOT, 'excel', 'ccisp'))